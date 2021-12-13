# Writing data to excel spreadsheet
from openpyxl import Workbook

# Instantiate a workbook object for storing data
wb = Workbook()

# Creating a sheet in the workbook
data = wb.create_sheet("Data")

# Initializes the workbook for data storage
def init_wb():

    # Start in the first column
    col = 1
    
    # Write test number
    data.cell(column=col, row=1).value = 'Coupon'
        
    # Next column
    col +=1 

    # Write test data
    data.cell(column=col, row=1).value = 'Expiration Date'
    
    # Next column
    col += 1

    # Write test result    
    data.cell(column=col, row=1).value = 'Email'

    # Next column
    col +=1 

    # Write test data
    data.cell(column=col, row=1).value = 'Password'
    
    # Next column
    col +=1 

    # Write test data
    data.cell(column=col, row=1).value = 'First Name'
        
    # Next column
    col +=1 

    # Write test data
    data.cell(column=col, row=1).value = 'Last Name'
    # wb.remove("sheet")
    wb.save("CHIPOTLE_COUPON_DATA.xlsx")
    return

# Records the test data in an Excel spreadsheet
def record_test(email, password, first_name, last_name):

    # Start in the first column
    col = 1
    
    # Write test number
    data.cell(column=col, row=(test_no + 1)).value = test_no
    
    # Next column
    col += 1

    # Write test result    
    data.cell(column=col, row=(test_no + 1)).value = result

    # Next column
    col +=1 

    # Write test data
    data.cell(column=col, row=(test_no + 1)).value = test_data

    # Saving the workbook
    wb.save("CHIPOTLE_COUPON_DATA.xlsx")
